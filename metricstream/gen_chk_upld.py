# ============================================================================
# Name: gen_chk_upld.py
#
# Description:
# Generate Excel upload template for Checklist from NCUA Questionnaire
# template.
#
# Assumptions:
# Require cx_Oracle, OPENPYXL, JDCAL, and ET_XMLFILE Python libraries.
# ============================================================================
# Revision History:
# Date       By        Comment
# ---------- --------- -----------------------------------------------------
# 03/21/2019 jhoang    Original release.
# 03/25/2019 jhoang    Added capability to query Oracle for Risk OBJ_ID.
# 04/19/2019 jhoang    Replace linefeed w/HTML break.
# 08/12/2019 jhoang    Query for both Risks and SubRisks.
# 08/13/2019 jhoang    Add exception handling.
# 09/12/2019 jhoang    Additional exception handling
# ============================================================================
 
import os
import getpass
import cx_Oracle
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from string import ascii_uppercase
 
ldiv   = '='.ljust (50, '=')
alpha  = list (ascii_uppercase)
 
tmpfil = "ncua_checklist.xlsx"
ResTyp = ['Yes or No', 'List of Responses', 'Text', 'Number', 'Amount', 'Date']
Risks  = ['Compliance', 'Credit', 'Interest Rate', 'Liquidity', 'Reputation'
         ,'Strategic', 'Transaction']
RiskId = {'Compliance':     'RISK-0000001006'
         ,'Credit':         'RISK-0000001005'
         ,'Interest Rate':  'RISK-0000001002'
         ,'Liquidity':      'RISK-0000001004'
         ,'Reputation':     'RISK-0000001008'
         ,'Strategic':      'RISK-0000001007'
         ,'Transaction':    'RISK-0000001003'}
 
 
# ====================================================================
# Add Section Name to List
# ====================================================================
 
def add_section (sdict, sec_name):
    if sec_name not in sdict.values():
        nsec = len (sdict)
        skey = 'SEC'+str(nsec).rjust (5,'0')
        sdict[skey] = sec_name
    else:
        for key in sdict.keys():
            if sdict[key] == sec_name:
                skey = key
    return skey
 
 
# ====================================================================
# Output Worksheet Header Row
# ====================================================================
 
def out_hdr (ws, hdr):
    col = 0
    bgc = PatternFill (fgColor="98FB98", fill_type="solid")
    for hval in hdr:
        ws[alpha[col]+'1'] = hval
        ws.column_dimensions[alpha[col]].width = 30
        col+=1
    for rows in ws.iter_rows (min_row=1, max_row=1, min_col=1):
        for cell in rows:
            cell.fill = bgc
 
 
# ====================================================================
# Open Excel Input workbook
# ====================================================================
 
tmp_wb = openpyxl.load_workbook (filename = tmpfil);
tmp_sheet = tmp_wb.active
 
while True:
    try:
        fname = input ("\nInput Excel Workbook Name: ")
        inp_wb = openpyxl.load_workbook (filename = fname);
    except FileNotFoundError:
        print ("Excel Workbook not found!")
    except:
        print (sys.exc_info()[0])
    else:
        break
 
# ====================================================================
# Obtain Basic Questionnaire info and copy to Uploader Template
# ====================================================================
 
inp_wb.active = 0
inp_sheet = inp_wb.active
cklist_name = inp_sheet.cell (row=5, column=5).value
cklist_inst = inp_sheet.cell (row=7, column=5).value
cklist_type = inp_sheet.cell (row=10, column=3).value
cklist_cont = inp_sheet.cell (row=10, column=5).value
cklist_risk = inp_sheet.cell (row=13, column=3).value
cklist_subr = inp_sheet.cell (row=15, column=3).value
instr_str   = cklist_inst.replace ('\n', '<br>')
 
tmp_sheet['C9']  = cklist_name.strip()
tmp_sheet['E9']  = instr_str.strip()
tmp_sheet['C12'] = cklist_type.strip()
tmp_sheet['E12'] = cklist_cont.strip()
 
# ====================================================================
# Loop through all Questions on 2nd Sheet and output to Uploader Temp
# ====================================================================
 
sec_pkey = None
sec_dict = {'key':'xx'}
qst_pkey = ['Q' for idex in range(1000)]
 
tmp_ws = tmp_wb["Question"]
inp_wb.active = 1
qst           = inp_wb.active
qst_rows      = qst.max_row
 
rnum = 1
for qrow in range (2,qst_rows+1):
    rnum += 1
    qst_num  = qst.cell (row=qrow, column=1).value
    section  = qst.cell (row=qrow, column=2).value
    question = qst.cell (row=qrow, column=3).value
    descrip  = qst.cell (row=qrow, column=4).value
    resptype = qst.cell (row=qrow, column=5).value
   lov_vals = qst.cell (row=qrow, column=6).value
    dep_qst  = qst.cell (row=qrow, column=7).value
    dep_val  = qst.cell (row=qrow, column=8).value
 
    if question == None:
        break
 
    qpk = 'QUE' + str(qst_num).rjust (5,'0')
    qst_pkey[qst_num] = qpk
 
    if section == None and sec_pkey == None:
        sec_pkey = 'General'
    if section == None:
        section = sec_pkey
    elif dep_qst == None:
        sec_pkey = add_section (sec_dict, section)
 
    #if question.split()[0].isdigit():
        #question = ' '.join (question.split()[1:])
    if resptype == None:
        resptype = 'Yes or No'
    elif resptype == 'List of Values':
        resptype = 'List of Responses'
    if descrip == None or len (descrip) == 0:
        descrip = 'Please respond below.'
 
    qst_str = question.replace ('\n', '  ')
    des_str = descrip.replace  ('\n', '<br>')
 
    tmp_ws['A'+str(rnum)] = cklist_name.strip()
    tmp_ws['B'+str(rnum)] = qpk
    tmp_ws['D'+str(rnum)] = qst_str.strip()
    tmp_ws['E'+str(rnum)] = des_str.strip()
    tmp_ws['H'+str(rnum)] = resptype.strip()
    tmp_ws['I'+str(rnum)] = dep_val
 
    if resptype == 'List of Responses':
        if lov_vals == None:
            tmp_ws['K'+str(rnum)].fill = PatternFill (fgColor="FAE8E4", fill_type="solid")
        else:
            tmp_ws['K'+str(rnum)] = lov_vals
 
    if dep_qst == None:
        tmp_ws['L'+str(rnum)] = sec_pkey
        tmp_ws['C'+str(rnum)] = 'No'
        tmp_ws['M'+str(rnum)] = 'SEC'
    else:
        tmp_ws['L'+str(rnum)] = qst_pkey[dep_qst]
        tmp_ws['C'+str(rnum)] = 'Yes'
        tmp_ws['M'+str(rnum)] = 'QST'
        for rows in tmp_ws.iter_rows (min_row=rnum, max_row=rnum, max_col=14):
            for cell in rows:
                cell.fill = PatternFill (fgColor="FFFFE0", fill_type="solid")
 
    if not resptype.strip() in ResTyp:
        tmp_ws['H'+str(rnum)].fill = PatternFill (fgColor="FAE8E4", fill_type="solid")
 
 
# ====================================================================
# Generate Sections Sheet
# ====================================================================
 
rnum = 1
snum = 0
tmp_ws = tmp_wb["Section"]
 
for sec_pk in sec_dict:
    if snum > 0:
        rnum += 1
        tmp_ws['A'+str(rnum)] = cklist_name
        tmp_ws['B'+str(rnum)] = sec_pk
        tmp_ws['C'+str(rnum)] = 'Yes'
        tmp_ws['D'+str(rnum)] = sec_dict[sec_pk]
    snum += 1
 
 
# ====================================================================
# Connect to Oracle
# ====================================================================
 
ouser = "ESS_CONV"
sid   = {'d': 'localhost:12000/ESSORAD'
        ,'t': 'localhost:11000/ESSORAD'
        ,'s': 'localhost:14000/ESSORAS'
        ,'p': 'localhost:10000/ESSORAP'}
 
while True:
    try:
        ienv = input ("\nEnv (Dev[d], Test[t], Stage[s], or Prod[p]")
        ipwd = getpass.getpass (prompt="ESS_CONV Password: ")
        env  = ienv[0].lower()
 
        con_str = ouser + "/" + ipwd + "@" + sid[env]
        con     = cx_Oracle.connect (con_str)
        cur     = con.cursor ()
    except cx_Oracle.DatabaseError:
        print ("Unable to connect to Oracle ("+ouser+"@"+sid[env]+")")
    except MyException at err:
        print (err)
    else:
        break
 
 
# ====================================================================
# Generate Risks Related To Tab
# ====================================================================
 
rnum = 2
tmp_ws = tmp_wb["Related to Tab"]
 
if not cklist_risk == None:
    tmp_ws['A'+str(rnum)] = cklist_name
    tmp_ws['B'+str(rnum)] = 'Risks'
    #obj_id = RiskId.get(cklist_risk)
    risk  = cklist_risk + ' Risk'
    sql   =  "select object_id from metricstream.ms_grc_risk " \
             " where upper(cast (object_name as varchar2(100))) = upper(:1) and object_level = 1"
    cur.execute (sql, (risk,))
    res = cur.fetchone ()
    if res == None:
        tmp_ws['C'+str(rnum)] = cklist_risk
        tmp_ws['C'+str(rnum)].fill = PatternFill (fgColor="FAE8E4", fill_type="solid")
    else:
        tmp_ws['C'+str(rnum)] = res[0]
    rnum += 1
    cur.close ()
    con.close ()
 
# ====================================================================
# If SubRisk provided, Connect to Oracle
# ====================================================================
 
if not cklist_subr == None:
 
    con   = cx_Oracle.connect (con_str)
    cur   = con.cursor ()
    risk  = cklist_risk + ' Risk'
    srisk = cklist_subr.strip()
    sql   = "select r.object_id from metricstream.ms_grc_risk r " \
             "inner join metricstream.ms_grc_risk_obj_mls_summary rel " \
             " on (cast (rel.cat_name_agg as varchar2(30)) = :1 and " \
             " rel.object_id = r.object_id) " \
             "where cast (r.object_name as varchar2(30)) = :2"
    cur.execute (sql, (risk, srisk))
    res = cur.fetchone ()
 
    tmp_ws['A'+str(rnum)] = cklist_name
    tmp_ws['B'+str(rnum)] = 'Risks'
    if res == None:
        tmp_ws['C'+str(rnum)] = cklist_subr
        tmp_ws['C'+str(rnum)].fill = PatternFill (fgColor="FAE8E4", fill_type="solid")
        print ("SubRisk "+cklist_subr+" not found")
    else:
        tmp_ws['C'+str(rnum)] = res[0]
 
    cur.close ()
    con.close ()
 
 
# ====================================================================
# Save New Workbook as Questionnaire + xlsx
# ====================================================================
 
print (cklist_name+".xlsx")
tmp_wb.save (cklist_name+".xlsx")
