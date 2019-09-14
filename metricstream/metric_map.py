# ============================================================================
# Name: merit_map.py
#
# Description:
# Search MERIT dictionary (tables/columns) for possible source of data.
#
# ============================================================================
# Revision History:
# Date       By        Comment
# ---------- --------- -----------------------------------------------------
# 08/20/2019 jhoang    Original release.
# ============================================================================
 
import os
import sys
import string
import getpass
import cx_Oracle
import openpyxl
import datetime
 
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.worksheet.table import Table, TableStyleInfo
from string import ascii_uppercase
 
sid = {'d': 'localhost:12000/ESSORAD'
      ,'t': 'localhost:11000/ESSORAD'
      ,'s': 'localhost:14000/ESSORAS'
      ,'p': 'localhost:10000/ESSORAP'
      }
pwd = {'d': 'ChangeMeJ', 't': 'ChangeMeJ', 's': 'ChangeMeJ', 'p': 'ChangeMeJ'}
dbe = {'d': 'DEV' ,'t': 'TEST' ,'s': 'STAGE' ,'p': 'PROD'}
now = datetime.datetime.now()
dtt = now.strftime ("%m/%d/%Y %H:%M:%S")
dts = now.strftime ("%Y%m%d_%H%M")
 
# ====================================================================
# Functions for Maintaining Tables/Columns
# ====================================================================
 
def add_tblist (lst, tbname):
    if tbname != None and tbname not in lst:
        lst.append (tbname)
 
def add_collist (lst, colname):
    if colname != None and colname not in lst:
        lst.append (colname)
 
 
 
# ====================================================================
# Prompt for Search String
# ====================================================================
 
istr = input ("\nSearch MERIT For: ")
lstr  = istr.lower()
 
# ====================================================================
# Prompt for Desired Database Environment
# ====================================================================
 
while True:
    try:
      # ipwd = getpass.getpass (prompt="ESS_CONV Password: ")
        ienv = input ("\nDatabase Environment (Dev[d], Test[t], Stage[s], or Prod[p]): ")
        env  = ienv[0].lower()
 
        cstr = "METRICSTREAM/" + pwd[env] + "@" + sid[env]
        con  = cx_Oracle.connect (cstr)
       cur  = con.cursor ()
    except cx_Oracle.DatabaseError:
        print ("Unable to connect to Oracle (metricstream@"+sid[env]+")")
    except:
        print (sys.exc_info() [0])
    else:
        break
 
 
rWb = Workbook ()
rws = rWb.active
bgc = PatternFill (fgColor="98FB98", fill_type="solid")
rws.title = "MERIT Search"
rws.sheet_properties.tabColor = "97AFE5"
 
tbx = rws['B1']
tbx.font = Font (color=colors.COLOR_INDEX[32], bold=True, size=12)
rws['B1'] = "Search For '" + istr + "' - Generated " + dtt
 
headers = ["Ref Object", "Label", "Form Column", "Table", "Column", "Data Type"
          ,"Validation Infolet" , "Vldn Info Parms", "Vldn Info Filter", "Vldn Filter Parms"
          ,"Default Infolet", "Def Info Parms", "Def Info Filter", "Def Filter Parms"
          ,"Dsply Infolet", "Config Extnd"]
rws.append (headers)
alp = list (string.ascii_uppercase)
for n in range (16):
    rws.column_dimensions[alp[n]].width = 30
rws.column_dimensions["D"].width = 35
rws.column_dimensions["E"].width = 35
rws.column_dimensions["F"].width = 15
rws.column_dimensions["P"].width = 15
 
# ====================================================================
# Query MERIT for Search String and output to Sheet
# ====================================================================
 
srch = " with obj as" \
"(select referenced_object, max(seq_no) mseqno " \
   " from metricstream.ms_apps_visual_entity" \
  " group by referenced_object)" \
" select obj.referenced_object" \
      " ,vea_title" \
      " ,result_column_name" \
      " ,decode(vea_multi_row_flag, 'N', vea_attr_source, 'Y', vea_attr_source || '_' || vea_region_name)" \
      " ,vea_attribute_id" \
      " ,case vea_datatype" \
          " when 3 then 'NUMBER ('||vea_field_size||')'" \
          " when 5 then 'VARCHAR2 ('||vea_field_size||')'" \
          " when 4 then 'DATE'" \
          " when 6 then 'CLOB'" \
       " end col_type" \
      " ,vea_validation_infolet" \
      " ,vea_validation_infolet_params" \
      " ,vea_validation_infolet_filter" \
      " ,cast (vea_validation_filter_params as varchar2(4000))" \
      " ,vea_default_infolet" \
      " ,vea_default_infolet_params" \
      " ,vea_default_infolet_filter" \
      " ,cast (vea_default_filter_params as varchar2(4000))" \
      " ,vea_display_infolet" \
      " ,config_extend_action" \
  " from metricstream.ms_apps_visual_entity_attr  frm" \
       " inner join obj on (obj.mseqno = frm.seq_no)" \
" where (lower (decode(vea_multi_row_flag, 'N', vea_attr_source, 'Y', vea_attr_source || '_' || vea_region_name)) like :txt" \
    " or lower (vea_attribute_id) like '%'||:txt||'%'" \
    " or lower (result_column_name) like '%'||:txt||'%'" \
    " or lower (vea_title) like '%'||:txt||'%')" \
" union " \
"select obj.referenced_object" \
     " ,utc.column_name as form_field_label" \
     " ,utc.column_name as form_column_name" \
     " ,vea_attr_source as table_name" \
     " ,utc.column_name as db_column_name" \
     " ,case" \
         " when utc.data_type like '%CHAR%' then data_type||' ('||utc.data_length||')'" \
         " when utc.data_type like 'NUMBER' then" \
            " case" \
               " when utc.data_precision is null then data_type" \
               " else data_type||'('||utc.data_precision||','||utc.data_scale||')'" \
               " end" \
         " else data_type" \
      " end col_type" \
     " ,null" \
     " ,null" \
     " ,null" \
     " ,null" \
     " ,null" \
     " ,null" \
     " ,null" \
     " ,null" \
     " ,null" \
    " ,null" \
" from ms_apps_visual_entity_attr  frm1" \
      " inner join user_tab_columns  utc on (utc.table_name = frm1.vea_attr_source)" \
      " inner join obj                   on (obj.mseqno     = frm1.seq_no)" \
      " where lower (utc.column_name) like '%'||:txt||'%'" \
  " and primary_key = 'Y'" \
" order by 1, 2"
 
ctr = 2
tblist  = []
collist = []
cur.execute (srch, txt=lstr)
for col in cur:
    ctr += 1
    rws.append (col)
    add_tblist  (tblist,  col[3])
   add_collist (collist, col[4])
 
 
# ====================================================================
# Format results as table
# ====================================================================
 
tab = Table (displayName="SrchTable", ref="A2:P"+str(ctr))
sty = TableStyleInfo (name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False
                     ,showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = sty
rws.add_table (tab)
rws.freeze_panes = "B3"
 
 
# ====================================================================
# Display Each Table Definition
# ====================================================================
 
hdr = ["Column Name", "Data Type", "Nullable"]
sql = "select tab.column_name" \
            ",case when data_type like '%CHAR%' then data_type||' ('||data_length||')'" \
                 " when data_type = 'NUMBER' then" \
                      " case when data_scale = 0 then 'INTEGER'" \
                           " when data_precision is null then data_type" \
                           " else data_type||' ('||data_precision||','||data_scale||')' end" \
                 " when data_type = 'FLOAT' then data_type||' ('||data_precision||')'" \
                 " else data_type end col_type" \
            ",decode (nullable, 'N','NOT NULL', null)" \
       " from all_tab_columns tab" \
      " where tab.owner = 'METRICSTREAM'" \
        " and tab.table_name = :tbn" \
      " order by tab.column_id"
 
for tb in tblist:
    print (tb)
    tWs = rWb.create_sheet (tb)
    tWs.title = tb
    tWs.append (hdr)
    tWs.column_dimensions['A'].width = 45
    tWs.column_dimensions['B'].width = 25
    tWs.column_dimensions['C'].width = 15
    cur.execute (sql, tbn=tb)
 
    ctr = 1
    for col in cur:
        ctr += 1
        tWs.append (col)
 
    if ctr == 1:
        tWs['A2'] = 'Table not found in METRICSTREAM schema'
        tWs.sheet_properties.tabColor = "E86370"
    else:
        tab = Table (displayName="Table"+tb, ref="A1:C"+str(ctr))
        sty = TableStyleInfo (name="TableStyleMedium7", showFirstColumn=False, showLastColumn=False
                             ,showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = sty
        tWs.add_table (tab)
        tWs.freeze_panes = "A2"
 
 
# ====================================================================
# Save New Workbook as MERIT Srch-[env]-[yyyymmdd].xlsx
# ====================================================================
 
rWb.save ("MERIT Srch-"+dbe[env]+"-"+dts+".xlsx")
print ("\nResults output to: MERIT Srch-"+dbe[env]+"-"+dts+".xlsx")
 
cur.close ()
con.close ()
